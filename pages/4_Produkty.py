import os
import math
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Produkty", page_icon="🍽️", layout="wide")

DATA_DIR = os.environ.get("DATA_DIR", "/data")
EXPORT_FILE = os.path.join(DATA_DIR, "export.xlsx")
EXPORT_SHEET = "export"


# =========================================================
# POMOCNÉ FUNKCE
# =========================================================
def clean_text(v):
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip()


def normalize_text(v):
    return clean_text(v).lower()


def normalize_colname(name):
    return clean_text(name).lower()


def is_blank(v):
    return clean_text(v) == ""


def number_or_blank(v):
    """
    Vrátí:
    - "" pokud je prázdno nebo 0
    - int pokud je celé číslo
    - float pokud je desetinné číslo
    """
    if v is None or v == "":
        return ""

    try:
        num = float(v)
    except Exception:
        return ""

    if math.isnan(num):
        return ""

    if num < 0:
        return ""

    if num == 0:
        return ""

    if num.is_integer():
        return int(num)

    return num


def first_existing_col(df, possible_names):
    norm_map = {normalize_colname(c): c for c in df.columns}
    for name in possible_names:
        if normalize_colname(name) in norm_map:
            return norm_map[normalize_colname(name)]
    return None


def unique_nonempty_from_column(df, col_name):
    if not col_name or col_name not in df.columns:
        return []

    vals = []
    for v in df[col_name].tolist():
        txt = clean_text(v)
        if txt:
            vals.append(txt)

    return sorted(set(vals), key=lambda x: x.lower())


def collect_ingredient_options(df):
    vals = []
    for c in df.columns:
        c_norm = normalize_colname(c)
        if "složení" in c_norm or "slozeni" in c_norm:
            for v in df[c].tolist():
                txt = clean_text(v)
                if txt:
                    vals.append(txt)

    return sorted(set(vals), key=lambda x: x.lower())


def resolve_value(new_value, selected_value):
    new_value = clean_text(new_value)
    selected_value = clean_text(selected_value)

    if new_value:
        return new_value
    if selected_value == "---":
        return ""
    return selected_value


def load_export_df():
    if not os.path.exists(EXPORT_FILE):
        return pd.DataFrame()

    return pd.read_excel(EXPORT_FILE, sheet_name=EXPORT_SHEET)


def get_next_id(df):
    if "ID" not in df.columns:
        return ""

    ids = pd.to_numeric(df["ID"], errors="coerce").dropna()
    if ids.empty:
        return 1

    return int(ids.max()) + 1


def append_product_to_export(new_row_dict):
    wb = load_workbook(EXPORT_FILE)

    if EXPORT_SHEET not in wb.sheetnames:
        raise ValueError(f"List '{EXPORT_SHEET}' neexistuje.")

    ws = wb[EXPORT_SHEET]
    headers = [cell.value for cell in ws[1]]

    row_values = [new_row_dict.get(h, "") for h in headers]
    ws.append(row_values)
    wb.save(EXPORT_FILE)


def column_matches_number_slot(col_norm, slot_num):
    """
    Pozná sloupce pro gramáže:
    např.
    - hmotnost 1
    - hmotnost 1 - suroviny ve sloupci l
    - 1 - suroviny ve sloupci ...
    """
    if "hmotnost" in col_norm and str(slot_num) in col_norm:
        return True

    # záložní logika kdyby někde chybělo slovo hmotnost
    if f"{slot_num} -" in col_norm and "suroviny" in col_norm:
        return True

    return False


def build_new_row(headers, df, values):
    row = {h: "" for h in headers}

    for h in headers:
        h_clean = normalize_colname(h)

        if h_clean == "id":
            row[h] = get_next_id(df)

        elif h_clean in ["název produktu", "nazev produktu"]:
            row[h] = clean_text(values["nazev_produktu"])

        elif h_clean == "kategorie":
            row[h] = clean_text(values["kategorie"])

        elif h_clean == "hmotnost":
            row[h] = number_or_blank(values["hmotnost"])

        elif h_clean == "velikost":
            row[h] = number_or_blank(values["velikost"])

        elif h_clean in ["základ", "zaklad"]:
            row[h] = clean_text(values["zaklad"])

        elif h_clean in ["počet kusů pečiva", "pocet kusu peciva", "počet ks", "pocet ks", "počet kusů", "pocet kusu"]:
            row[h] = number_or_blank(values["pocet_ks"])

        elif h_clean in ["mazání", "mazani"]:
            row[h] = clean_text(values["mazani"])

        elif "hmotnost suroviny ve sloupci j" in h_clean or h_clean in ["hmotnost mazání", "hmotnost mazani"]:
            row[h] = number_or_blank(values["hmotnost_mazani"])

        elif h_clean in ["složení 1", "slozeni 1"]:
            row[h] = clean_text(values["slozeni_1"])
        elif column_matches_number_slot(h_clean, 1):
            row[h] = number_or_blank(values["hmotnost_1"])

        elif h_clean in ["složení 2", "slozeni 2"]:
            row[h] = clean_text(values["slozeni_2"])
        elif column_matches_number_slot(h_clean, 2):
            row[h] = number_or_blank(values["hmotnost_2"])

        elif h_clean in ["složení 3", "slozeni 3"]:
            row[h] = clean_text(values["slozeni_3"])
        elif column_matches_number_slot(h_clean, 3):
            row[h] = number_or_blank(values["hmotnost_3"])

        elif h_clean in ["složení 4", "slozeni 4"]:
            row[h] = clean_text(values["slozeni_4"])
        elif column_matches_number_slot(h_clean, 4):
            row[h] = number_or_blank(values["hmotnost_4"])

        elif h_clean in ["složení 5", "slozeni 5"]:
            row[h] = clean_text(values["slozeni_5"])
        elif column_matches_number_slot(h_clean, 5):
            row[h] = number_or_blank(values["hmotnost_5"])

    return row


def validate_form(
    export_df,
    col_nazev,
    nazev_produktu,
    hmotnost,
    velikost,
    pocet_ks,
    hmotnost_mazani,
    slozeni_hmotnosti,
):
    errors = []

    # povinný název
    if is_blank(nazev_produktu):
        errors.append("Vyplň název produktu.")

    # duplicita názvu
    if col_nazev and col_nazev in export_df.columns and not is_blank(nazev_produktu):
        dup = (
            export_df[col_nazev]
            .astype(str)
            .str.strip()
            .str.lower()
            .eq(clean_text(nazev_produktu).lower())
            .any()
        )
        if dup:
            errors.append("Takový produkt už v exportu existuje.")

    # čísla nesmí být záporná
    numeric_fields = [
        ("Hmotnost", hmotnost),
        ("Velikost", velikost),
        ("Počet kusů pečiva", pocet_ks),
        ("Hmotnost mazání", hmotnost_mazani),
    ]

    for label, value in numeric_fields:
        if value is not None and value < 0:
            errors.append(f"{label} nesmí být záporná.")

    # kontrola složení + gramáže
    for idx, (slozeni, gram) in enumerate(slozeni_hmotnosti, start=1):
        slozeni_txt = clean_text(slozeni)

        if slozeni_txt and (gram is None or gram <= 0):
            errors.append(f"U 'Složení {idx}' chybí hmotnost nebo je 0.")

        if not slozeni_txt and gram is not None and gram > 0:
            errors.append(f"U 'Hmotnost {idx}' chybí surovina.")

        if gram is not None and gram < 0:
            errors.append(f"Hmotnost {idx} nesmí být záporná.")

    return errors


# =========================================================
# UI
# =========================================================
st.title("Produkty")
st.caption("Přidávání nových produktů do listu export")

try:
    export_df = load_export_df()
except Exception as e:
    st.error(f"Chyba při načtení exportu: {e}")
    st.stop()

if export_df.empty:
    st.error("Soubor export.xlsx nebo list 'export' nebyl nalezen nebo je prázdný.")
    st.stop()

headers = list(export_df.columns)

# důležité sloupce
col_nazev = first_existing_col(export_df, ["Název produktu", "Nazev produktu"])
col_kategorie = first_existing_col(export_df, ["kategorie", "Kategorie"])
col_zaklad = first_existing_col(export_df, ["základ", "zaklad", "Základ"])
col_mazani = first_existing_col(export_df, ["mazání", "mazani", "Mazání", "Mazani"])

# dropdown možnosti
default_categories = [
    "Bagety",
    "Chlebíčky",
    "Saláty",
    "Mísy a sety",
    "Mísy a setySaláty",
    "Sladké",
    "Dezerty",
    "Cukrárna",
    "Nápoje",
    "Nádobí",
    "Kanapky",
    "McKinsey",
    "Ostatní",
]

categories_existing = unique_nonempty_from_column(export_df, col_kategorie)
category_options = sorted(set(default_categories + categories_existing), key=lambda x: x.lower())

zaklad_options = unique_nonempty_from_column(export_df, col_zaklad)
mazani_options = unique_nonempty_from_column(export_df, col_mazani)
ingredient_options = collect_ingredient_options(export_df)

tab1, tab2 = st.tabs(["Seznam produktů", "Přidat produkt"])

with tab1:
    st.subheader("Aktuální produkty")

    search_text = st.text_input("Hledat produkt")
    filtered_df = export_df.copy()

    if search_text and col_nazev and col_nazev in filtered_df.columns:
        filtered_df = filtered_df[
            filtered_df[col_nazev].astype(str).str.contains(search_text, case=False, na=False)
        ]

    cols_to_show = [c for c in ["ID", "Název produktu", "kategorie", "hmotnost", "velikost"] if c in filtered_df.columns]

    if cols_to_show:
        st.dataframe(filtered_df[cols_to_show], use_container_width=True, height=500)
    else:
        st.dataframe(filtered_df, use_container_width=True, height=500)

with tab2:
    st.subheader("Nový produkt")

    with st.form("novy_produkt_form", clear_on_submit=True):
        c1, c2 = st.columns(2)

        with c1:
            nazev_produktu = st.text_input("Název produktu *").strip()

            selected_kategorie = st.selectbox("Kategorie - vyber", ["---"] + category_options)
            nova_kategorie = st.text_input("Kategorie - nebo napiš novou").strip()

            hmotnost = st.number_input("Hmotnost", min_value=0.0, step=1.0, format="%g")
            velikost = st.number_input("Velikost", min_value=0.0, step=1.0, format="%g")

            st.markdown("### Základ")
            selected_zaklad = st.selectbox("Základ - vyber", ["---"] + zaklad_options)
            novy_zaklad = st.text_input("Základ - nebo napiš nový").strip()

            pocet_ks = st.number_input("Počet kusů pečiva", min_value=0.0, step=1.0, format="%g")

            selected_mazani = st.selectbox("Mazání - vyber", ["---"] + mazani_options)
            nove_mazani = st.text_input("Mazání - nebo napiš nové").strip()

            hmotnost_mazani = st.number_input("Hmotnost mazání", min_value=0.0, step=1.0, format="%g")

        with c2:
            st.markdown("### Složení")

            sel_slozeni_1 = st.selectbox("Složení 1 - vyber", ["---"] + ingredient_options)
            new_slozeni_1 = st.text_input("Složení 1 - nebo napiš nové").strip()
            hmotnost_1 = st.number_input("Hmotnost 1", min_value=0.0, step=1.0, format="%g")

            sel_slozeni_2 = st.selectbox("Složení 2 - vyber", ["---"] + ingredient_options)
            new_slozeni_2 = st.text_input("Složení 2 - nebo napiš nové").strip()
            hmotnost_2 = st.number_input("Hmotnost 2", min_value=0.0, step=1.0, format="%g")

            sel_slozeni_3 = st.selectbox("Složení 3 - vyber", ["---"] + ingredient_options)
            new_slozeni_3 = st.text_input("Složení 3 - nebo napiš nové").strip()
            hmotnost_3 = st.number_input("Hmotnost 3", min_value=0.0, step=1.0, format="%g")

            sel_slozeni_4 = st.selectbox("Složení 4 - vyber", ["---"] + ingredient_options)
            new_slozeni_4 = st.text_input("Složení 4 - nebo napiš nové").strip()
            hmotnost_4 = st.number_input("Hmotnost 4", min_value=0.0, step=1.0, format="%g")

            sel_slozeni_5 = st.selectbox("Složení 5 - vyber", ["---"] + ingredient_options)
            new_slozeni_5 = st.text_input("Složení 5 - nebo napiš nové").strip()
            hmotnost_5 = st.number_input("Hmotnost 5", min_value=0.0, step=1.0, format="%g")

        ulozit = st.form_submit_button("Uložit nový produkt")

    if ulozit:
        final_kategorie = resolve_value(nova_kategorie, selected_kategorie)
        final_zaklad = resolve_value(novy_zaklad, selected_zaklad)
        final_mazani = resolve_value(nove_mazani, selected_mazani)

        final_slozeni_1 = resolve_value(new_slozeni_1, sel_slozeni_1)
        final_slozeni_2 = resolve_value(new_slozeni_2, sel_slozeni_2)
        final_slozeni_3 = resolve_value(new_slozeni_3, sel_slozeni_3)
        final_slozeni_4 = resolve_value(new_slozeni_4, sel_slozeni_4)
        final_slozeni_5 = resolve_value(new_slozeni_5, sel_slozeni_5)

        slozeni_hmotnosti = [
            (final_slozeni_1, hmotnost_1),
            (final_slozeni_2, hmotnost_2),
            (final_slozeni_3, hmotnost_3),
            (final_slozeni_4, hmotnost_4),
            (final_slozeni_5, hmotnost_5),
        ]

        errors = validate_form(
            export_df=export_df,
            col_nazev=col_nazev,
            nazev_produktu=nazev_produktu,
            hmotnost=hmotnost,
            velikost=velikost,
            pocet_ks=pocet_ks,
            hmotnost_mazani=hmotnost_mazani,
            slozeni_hmotnosti=slozeni_hmotnosti,
        )

        if errors:
            for err in errors:
                st.error(err)
            st.stop()

        values = {
            "nazev_produktu": nazev_produktu,
            "kategorie": final_kategorie,
            "hmotnost": hmotnost,
            "velikost": velikost,
            "zaklad": final_zaklad,
            "pocet_ks": pocet_ks,
            "mazani": final_mazani,
            "hmotnost_mazani": hmotnost_mazani,
            "slozeni_1": final_slozeni_1,
            "hmotnost_1": hmotnost_1,
            "slozeni_2": final_slozeni_2,
            "hmotnost_2": hmotnost_2,
            "slozeni_3": final_slozeni_3,
            "hmotnost_3": hmotnost_3,
            "slozeni_4": final_slozeni_4,
            "hmotnost_4": hmotnost_4,
            "slozeni_5": final_slozeni_5,
            "hmotnost_5": hmotnost_5,
        }

        try:
            new_row = build_new_row(headers, export_df, values)
            append_product_to_export(new_row)
            st.success(f"Produkt '{nazev_produktu}' byl přidán.")
            st.rerun()
        except Exception as e:
            st.error(f"Uložení selhalo: {e}")
